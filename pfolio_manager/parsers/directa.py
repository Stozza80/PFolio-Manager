"""Parser for Directa SIM "P_TOTALE" xlsx exports."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .. import util
from ..models import ParsedLine, ParsedReport
from . import ParserError, find_header_row, header_name_to_col, to_float

HEADER_SIGNATURE = {"Strumento", "Ticker", "Isin"}


def parse(file_path: Path) -> ParsedReport:
    # read_only=False: some broker exports declare an incorrect sheet dimension,
    # which breaks openpyxl's read-only row iteration (confirmed with BNL's export).
    wb = load_workbook(file_path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    account_id = None
    as_of_date = None
    total_value_native = None
    for row in rows[:8]:
        text = str(row[0]) if row and row[0] is not None else ""
        if text.startswith("Conto"):
            match = re.search(r"Conto\s*:\s*(\S+)", text)
            if match:
                account_id = match.group(1)
        elif text.startswith("Data estrazione"):
            as_of_date = util.parse_directa_extraction_date(text).date().isoformat()
        elif text.startswith("Valore portafoglio"):
            total_value_native = util.parse_it_number(text)

    if account_id is None:
        raise ParserError("Directa file: could not find 'Conto' metadata row")
    if as_of_date is None:
        raise ParserError("Directa file: could not find 'Data estrazione' metadata row")

    header_idx = find_header_row(rows, HEADER_SIGNATURE)
    cols = header_name_to_col(rows[header_idx])

    lines: list[ParsedLine] = []
    for row in rows[header_idx + 1 :]:
        isin = row[cols["Isin"]]
        quantity = to_float(row[cols["Quantita"]])
        if not isin or quantity is None:
            break  # totals row / trailing blank row

        lines.append(
            ParsedLine(
                isin=str(isin).strip(),
                instrument_name=str(row[cols["Strumento"]]).strip(),
                ticker=str(row[cols["Ticker"]]).strip() if row[cols["Ticker"]] else None,
                market=None,
                currency=str(row[cols["Divisa"]]).strip() if row[cols["Divisa"]] else "EUR",
                quantity=quantity,
                avg_cost_price=to_float(row[cols["Prezzo medio"]]),
                cost_value=to_float(row[cols["Valore di carico"]]),
                market_price=to_float(row[cols["Prezzo"]]),
                market_value=to_float(row[cols["Valore attuale"]]),
                unrealized_gain_loss=to_float(row[cols["Gain/Loss €"]]),
                unrealized_gain_loss_pct=to_float(row[cols["Gain/Loss %"]]),
                accrued_interest=None,
                broker_fx_rate=None,  # Directa exports have no FX-rate column
            )
        )

    return ParsedReport(
        broker="directa",
        account_id=account_id,
        as_of_date=as_of_date,
        currency="EUR",
        total_value_native=total_value_native,
        source_file=file_path.name,
        lines=lines,
    )
