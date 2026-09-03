"""Parser for BNL Trading "PosizioneDep" xlsx exports."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .. import util
from ..models import ParsedLine, ParsedReport
from . import ParserError, find_header_row, header_name_to_col, to_float

HEADER_SIGNATURE = {"Titolo", "ISIN", "Cambio Corrente"}
METADATA_LABEL = "Situazione del"


def parse(file_path: Path) -> ParsedReport:
    # read_only=False: this broker's export declares an incorrect sheet dimension
    # (<dimension ref="A1"/> despite having 25 rows), which makes openpyxl's
    # read-only row iteration silently return a single empty row.
    wb = load_workbook(file_path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    label_row_idx = None
    for idx, row in enumerate(rows[:10]):
        texts = {str(c).strip() for c in row if isinstance(c, str)}
        if METADATA_LABEL in texts:
            label_row_idx = idx
            break
    if label_row_idx is None:
        raise ParserError(f"BNL file: metadata label {METADATA_LABEL!r} not found")

    labels = header_name_to_col(rows[label_row_idx])
    values = rows[label_row_idx + 1]

    account_id = str(values[labels["Deposito numero"]]).strip()
    situazione = values[labels[METADATA_LABEL]]
    # openpyxl resolves the Excel date serial to a datetime automatically in most
    # cases; fall back to manual conversion if a raw serial number slips through.
    situazione_dt = situazione if isinstance(situazione, datetime) else util.excel_serial_to_datetime(float(situazione))
    as_of_date = situazione_dt.date().isoformat()
    total_value_native = to_float(values[labels["Controvalore*"]])

    header_idx = find_header_row(rows, HEADER_SIGNATURE)
    cols = header_name_to_col(rows[header_idx])

    lines: list[ParsedLine] = []
    for row in rows[header_idx + 1 :]:
        isin = row[cols["ISIN"]]
        quantity = to_float(row[cols["Quantita'"]])
        if not isin or quantity is None:
            break  # totals row / disclaimer row / trailing blank row

        lines.append(
            ParsedLine(
                isin=str(isin).strip(),
                instrument_name=str(row[cols["Titolo"]]).strip(),
                ticker=None,  # BNL exports don't include a ticker/symbol column
                market=str(row[cols["Mercato"]]).strip() if row[cols["Mercato"]] else None,
                currency=str(row[cols["Valuta"]]).strip() if row[cols["Valuta"]] else "EUR",
                quantity=quantity,
                avg_cost_price=to_float(row[cols["P.zo medio di carico"]]),
                cost_value=to_float(row[cols["Controvalore di carico"]]),
                market_price=to_float(row[cols["P.zo di mercato"]]),
                market_value=to_float(row[cols["Valore Corrente"]]),
                unrealized_gain_loss=to_float(row[cols["Profit e loss"]]),
                unrealized_gain_loss_pct=to_float(row[cols["Profit e loss %"]]),
                accrued_interest=to_float(row[cols["Rateo"]]),
                broker_fx_rate=to_float(row[cols["Cambio Corrente"]]),
            )
        )

    return ParsedReport(
        broker="bnl",
        account_id=account_id,
        as_of_date=as_of_date,
        currency="EUR",
        total_value_native=total_value_native,
        source_file=file_path.name,
        lines=lines,
    )
