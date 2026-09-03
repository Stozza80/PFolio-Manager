"""Structural broker detection.

Deliberately does NOT use filenames (account numbers / dates vary run to run) —
instead matches file extension + header/title signatures against known broker formats.
Never guesses: an unrecognized file raises DetectionError rather than being silently skipped.
"""
from __future__ import annotations

from pathlib import Path

import xlrd
from openpyxl import load_workbook

from .parsers import bnl, directa, fineco


class DetectionError(Exception):
    pass


def detect_broker(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".xls":
        return _detect_legacy_xls(file_path)
    if suffix == ".xlsx":
        return _detect_xlsx(file_path)
    raise DetectionError(
        f"Unrecognized file extension for {file_path.name!r}: only .xls/.xlsx are supported"
    )


def _detect_legacy_xls(file_path: Path) -> str:
    # Fineco is currently the only broker using the legacy binary .xls format.
    wb = xlrd.open_workbook(file_path)
    sh = wb.sheet_by_index(0)
    first_rows = [sh.row_values(r) for r in range(min(sh.nrows, 5))]
    if any(fineco.TITLE_MARKER in str(cell) for row in first_rows for cell in row):
        return "fineco"
    raise DetectionError(
        f"{file_path.name!r} is a legacy .xls but doesn't look like a Fineco export "
        f"(missing {fineco.TITLE_MARKER!r} in the first rows)"
    )


def _detect_xlsx(file_path: Path) -> str:
    # read_only=False: see the comment in parsers/bnl.py — some broker exports declare
    # an incorrect sheet dimension that breaks openpyxl's read-only row iteration.
    wb = load_workbook(file_path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    for row in rows[:15]:
        texts = {str(c).strip() for c in row if isinstance(c, str) and str(c).strip()}
        if directa.HEADER_SIGNATURE.issubset(texts):
            return "directa"
        if bnl.HEADER_SIGNATURE.issubset(texts):
            return "bnl"
    raise DetectionError(
        f"{file_path.name!r} is an .xlsx but its header doesn't match any known broker "
        f"format (Directa/BNL)"
    )
